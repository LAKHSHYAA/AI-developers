from openpyxl import Workbook, load_workbook

# Create a new Excel workbook
wb = Workbook()

# Create a new sheet
ws = wb.active
ws.title = "Data"

# Define data
data = [
    ["Name", "Age", "City"],
    ["Alice", 25, "New York"],
    ["Bob", 30, "San Francisco"],
    ["Charlie", 35, "Chicago"],["lkah",19,"salmem"]
]

# Insert data into the sheet
for row in data:
    ws.append(row)

# Save the workbook
wb.save("example.xlsx")

# Read the Excel file
wb = load_workbook("example.xlsx")
ws = wb.active

# Print the data
print("Reading data from Excel:")
for row in ws.iter_rows(values_only=True):
    print(row)

# Update the Excel file
ws["B2"] = 26
wb.save("example.xlsx")

# Read the updated data
wb = load_workbook("example.xlsx")
ws = wb.active

# Print the updated data
print("\nReading updated data from Excel:")
for row in ws.iter_rows(values_only=True):
    print(row)

# Delete a row from the Excel file
ws.delete_rows(2)
wb.save("example.xlsx")

# Read the data after deletion
wb = load_workbook("example.xlsx")
ws = wb.active

# Print the data after deletion
print("\nReading data after deletion from Excel:")
for row in ws.iter_rows(values_only=True):
    print(row)
