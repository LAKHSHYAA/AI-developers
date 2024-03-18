from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import numpy as np

# Create a new Excel workbook
wb = Workbook()

# Create a new sheet
ws = wb.active
ws.title = "Data"

# Define data
data = [
    ["Name", "Age", "City", "Score"],
    ["Alice", 25, "New York", 85],
    ["Bob", 30, "San Francisco", 92],
    ["Charlie", 35, "Chicago", 78],
    ["David", 40, "Los Angeles", 95],
    ["Eve", 45, "Miami", 87]
]

# Insert data into the sheet
for row in data:
    ws.append(row)

# Save the workbook
wb.save("example.xlsx")

# Read the Excel file
wb = load_workbook("example.xlsx")
ws = wb.active

# Print the data
print("Reading data from Excel:")
for row in ws.iter_rows(values_only=True):
    print(row)

# Data visualization
names = [row[0] for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1, values_only=True)]
ages = [row[1] for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2, values_only=True)]
scores = [row[3] for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4, values_only=True)]

# Bar chart for age distribution
plt.figure(figsize=(10, 4))
plt.subplot(1, 2, 1)
plt.bar(names, ages, color='skyblue')
plt.xlabel('Name')
plt.ylabel('Age')
plt.title('Age Distribution')
plt.xticks(rotation=45)

# Pie chart for score distribution
plt.subplot(1, 2, 2)
explode = (0.1, 0, 0, 0, 0)
plt.pie(scores, labels=names, autopct='%1.1f%%', startangle=140, explode=explode)
plt.axis('equal')
plt.title('Score Distribution')

plt.tight_layout()
plt.show()
